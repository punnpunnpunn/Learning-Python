from openpyxl import load_workbook, workbook

workbook = load_workbook(filename="2D_list.xlsx")

workbook.sheetnames
sheet = workbook.active
sheet.title

matrix=[]
for y in range(10):
    list=[]
    for x in range(10):
        list.append(x+(y*10))
    matrix[y*10:(y*10)+9]=[list]

alpha=["A","B","C","D","E","F","G","H","I","J"]
for x in range (10):
    list=matrix[x]
    for y in range(10):
        a=str(list[y])
        sheet[alpha[x]+str(y+1)]=a
workbook.save(filename="2D_list.xlsx")