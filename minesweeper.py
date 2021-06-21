from openpyxl import load_workbook
import random
workbook = load_workbook(filename="minesweeper.xlsx")
workbook.sheetnames

sheet = workbook.active
sheet

sheet.title

char = ["A","B","C","D","E","F","G","H","I","J"]

for i in range(0,10):                  
    for j in range(1,11):
        sheet[char[i]+str(j)] = " "

for x in range(1,11):
    sheet[char[random.randint(0,9)]+str(x)]="x"

workbook.save(filename="minesweeper.xlsx")

def mines():
    y=int(input("How many players?"))
    z=1
    x=1
    while x!=0:
        while z<=y:
            x=input("Player"+str(z))
            if sheet[x].value=="x":
                print("Game Over")
                x=0
                return()
            z=z+1
            if z>y:
                z=z-y
mines()