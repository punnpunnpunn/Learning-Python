from openpyxl import load_workbook, workbook

workbook = load_workbook(filename="score.xlsx")

workbook.sheetnames
sheet = workbook.active
sheet.title

print("Hi")
a=["B","C","D","E","F","G"]
for i in range (5,25):
    y=0
    for x in range(6):
        y=y+(sheet[a[x]+str(i)].value)
    sheet["I"+str(i)]=y
    if sheet["I"+str(i)].value > 80:
        sheet["J"+str(i)].value="A"
    elif sheet["I"+str(i)].value > 70:
        sheet["J"+str(i)].value="B"
    elif sheet["I"+str(i)].value > 60:
        sheet["J"+str(i)].value="C"
    elif sheet["I"+str(i)].value > 50:
        sheet["J"+str(i)].value="D"
    elif sheet["I"+str(i)].value <= 50:
        sheet["J"+str(i)].value="F"
workbook.save(filename="score.xlsx")